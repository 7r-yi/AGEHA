import openpyxl
from openpyxl.styles import Font
import re
import cv2
from PIL import Image
from pdf2image import convert_from_path
import numpy as np
import os
import win32com.client

flag_war_start = False
import_track = []
clan_war_time = 0


def maketable_introduction(flag):
    msg = "集計表を作成するお＾ｑ＾(Cancelで作成中止)\n" \
          "```自チームのメンバー [ふつきん,P2,P3,P4,P5,P6]\n自チームの各得点 [80,80,80,80,80,80]\n\n" \
          "敵チーム名 [ABC]\n対戦回数 [1]\n" \
          "敵チームのメンバー [P7,P8,P9,P10,P11,P12]\n敵チームの各得点 [80,80,80,80,80,80]"

    if flag == 1:
        return msg + "```※コースと交流戦日付は記入不要だお＾ｑ＾"
    else:
        return msg + "\n\nコース名 [T1,T2,T3,T4,T5,T6,T7,T8,T9,T10,T11,T12]\n"\
                     "自チームの選択コース [1,2,3,4,5,6]\n\n日付 [2020/01/01]```"


def maketable_check_input(str, flag):
    safe, msg = True, ""
    n1, n2 = str.count('['), str.count(']')

    if not str.startswith("自チームのメンバー"):
        safe = False
    elif n1 >= 6 and n2 >= 6:
        input = str.replace("[", "]").split("]")
        for i in range(1, 12, 2):
            if (i != 5 and i != 7) and input[i].count(',') != 5:
                safe = False
                msg = "名前または得点が6つずつ入力されていないかも？＾ｑ＾"
        if flag == 0 and safe:
            if not input[13].count(',') == 11:
                safe = False
            msg = "コースが12個入力されていないかも？＾ｑ＾"
        if flag == 1 and safe:
            if not n1 == 6 and n2 == 6:
                safe = False
            msg = "余計なデータがある or []で閉じられていない箇所があるかも？＾ｑ＾"
        elif flag == 0 and safe:
            if not n1 == 9 and n2 == 9:
                safe = False
            msg = "余計なデータがある or []で閉じられていない箇所があるかも？＾ｑ＾"
    else:
        safe = False
        msg = "[]で閉じられていない箇所があるかも？＾ｑ＾"

    return safe, msg


def make_table(input, flag):
    source_excel_pass = 'clanwar_table.xlsx'
    created_excel_pass = 'created_sheet.xlsx'
    pdf_pass = 'created_sheet.pdf'
    table_image_pass = 'table.png'
    bg_image_pass = 'clanwar_table_bg.png'
    created_image_pass = 'created_sheet.png'

    book = openpyxl.load_workbook(source_excel_pass)  # 集計表を作成
    sheet = book['6 vs 6']
    input = input.replace("[", "]").split("]")

    ally_name = input[1].split(',')
    for i in range(6):
        sheet.cell(3 + i, 4, ally_name[i])

    ally_point = input[3].split(',')
    for i in range(6):
        sheet.cell(3 + i, 6, int(ally_point[i]))

    sheet.cell(9, 1, input[5])

    enemy_name = input[9].split(',')
    for i in range(6):
        sheet.cell(9 + i, 4, enemy_name[i])

    enemy_point = input[11].split(',')
    for i in range(6):
        sheet.cell(9 + i, 6, int(enemy_point[i]))

    if flag == 1:
        for i in range(12):
            if "__" in import_track[i]:
                sheet.cell(3 + i, 12, re.sub('[_*]', "", import_track[i]))
                cell = sheet.cell(row=3 + i, column=12).coordinate
                sheet[cell].font = Font(color='00B0F0', size=26, name='Cataneo BT')
            else:
                sheet.cell(3 + i, 12, import_track[i])
        global clan_war_time
        date = clan_war_time
    else:
        track = input[13].split(',')
        select_num = input[15].split(',')
        for i in range(12):
            sheet.cell(3 + i, 12, track[i])
            for j in range(len(select_num)):
                if i + 1 == int(select_num[j]):
                    cell = sheet.cell(row=3 + i, column=12).coordinate
                    sheet[cell].font = Font(color='00B0F0', size=26, name='Cataneo BT')
        date = input[17]

    if int(input[7]) == 1:
        send_str = f"vs {input[5]} - {date}"
    else:
        send_str = f"vs {input[5]} ({input[7]}回目) - {date}"

    book.save(created_excel_pass)
    book.close()

    excel = win32com.client.Dispatch("Excel.Application")
    file = excel.Workbooks.Open(os.path.abspath(created_excel_pass))
    file.WorkSheets('6 vs 6').Select()
    file.ActiveSheet.ExportAsFixedFormat(0, os.path.abspath(pdf_pass))
    excel.Quit()

    pdf = convert_from_path(os.path.abspath(pdf_pass))  # pdfを画像化
    pdf[0].save(table_image_pass, "png")

    img = cv2.imread(table_image_pass)  # 余白を削除
    height, width = img.shape[:2]
    s = 1000
    h1, h2, w1, w2 = 0, 0, 0, 0

    for y in range(height):
        if all(img[y, s] != np.array([255, 255, 255])):
            h1 = y + 3
            break

    for y in reversed(range(height)):
        if all(img[y, s] != np.array([255, 255, 255])):
            h2 = y
            break

    for x in range(width):
        if all(img[s, x] != np.array([255, 255, 255])):
            w1 = x + 3
            break

    for x in reversed(range(width)):
        if all(img[s, x] != np.array([255, 255, 255])):
            w2 = x
            break

    crop_img = img[h1:h2, w1:w2]
    height, width = crop_img.shape[:2]
    crop_img = cv2.resize(crop_img, (int(width * 0.66), int(height * 0.66)))

    alpha_img = cv2.cvtColor(crop_img, cv2.COLOR_RGB2RGBA)  # 集計表の黒色を透過
    alpha_img[..., 3] = np.where(np.all(crop_img == 0, axis=-1), 0, 255)
    cv2.imwrite(table_image_pass, alpha_img)

    img1 = Image.open(bg_image_pass)  # 背景画像と集計表を合成
    img2 = Image.open(table_image_pass)
    img1.paste(img2, (0, 0), img2)
    img1.save(created_image_pass)

    os.remove(created_excel_pass)
    os.remove(pdf_pass)
    os.remove(table_image_pass)

    return send_str
